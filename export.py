import os
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DB_FILE = "sms_monitor.db"
KYIV_TIMEZONE = ZoneInfo("Europe/Kyiv")


def format_kyiv_time(value):
    if not value:
        return value

    try:
        parsed = datetime.fromisoformat(
            str(value).replace("Z", "+00:00")
        )
    except ValueError:
        return value

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    return parsed.astimezone(KYIV_TIMEZONE).strftime(
        "%Y-%m-%d %H:%M:%S %Z"
    )


# =========================================================
# DATABASE
# =========================================================

def get_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


# =========================================================
# EXTENSIONS
# =========================================================

def get_extensions():

    conn = get_connection()

    rows = conn.execute("""
        SELECT
            extension_id,
            extension_number,
            name,
            type
        FROM extensions
        WHERE active = 1
          AND type = 'User'
        ORDER BY extension_number
    """).fetchall()

    conn.close()

    return rows


# =========================================================
# MESSAGE COUNT
# =========================================================

def get_message_count(extension_id):

    conn = get_connection()

    row = conn.execute("""
        SELECT COUNT(*) AS total
        FROM messages
        WHERE extension_id = ?
    """, (extension_id,)).fetchone()

    conn.close()

    return row["total"]


# =========================================================
# WEB EXPORT
# =========================================================

def get_messages_by_extensions(
    extension_ids,
    date_from=None,
    date_to=None
):
    """
    Получает сообщения для web API export.

    extension_ids:
        [123, 456, 789]

    date_from:
        YYYY-MM-DD

    date_to:
        YYYY-MM-DD
    """

    if not extension_ids:
        return []

    conn = get_connection()

    placeholders = ",".join(
        ["?"] * len(extension_ids)
    )

    query = f"""
        SELECT
            m.ringcentral_id,
            m.extension_id,
            e.extension_number,
            e.name AS extension_name,

            m.from_number,
            m.to_number,
            m.direction,
            m.message,
            m.status,
            m.creation_time,
            m.delivery_time,
            m.last_updated

        FROM messages m

        LEFT JOIN extensions e
            ON e.extension_id = m.extension_id

        WHERE m.extension_id IN ({placeholders})
    """

    params = list(extension_ids)

    # -----------------------------------------------------
    # DATE FROM
    # -----------------------------------------------------

    if date_from:
        if "T" in date_from:
            query += """
                AND datetime(m.creation_time) >= datetime(?)
            """
        else:
            query += """
                AND date(m.creation_time) >= date(?)
            """

        params.append(date_from)

    # -----------------------------------------------------
    # DATE TO
    # -----------------------------------------------------

    if date_to:
        if "T" in date_to:
            query += """
                AND datetime(m.creation_time) <= datetime(?)
            """
        else:
            query += """
                AND date(m.creation_time) <= date(?)
            """

        params.append(date_to)

    # -----------------------------------------------------
    # ORDER
    # -----------------------------------------------------

    query += """
        ORDER BY
            m.creation_time ASC,
            m.id ASC
    """

    rows = conn.execute(
        query,
        params
    ).fetchall()

    conn.close()

    return rows


# =========================================================
# EXCEL GENERATOR
# =========================================================

def generate_excel_from_messages(
    rows,
    title="SMS Export"
):
    """
    Создаёт XLSX в памяти и возвращает BytesIO.
    """

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "SMS"

    # -----------------------------------------------------
    # HEADERS
    # -----------------------------------------------------

    headers = [
        "Extension",
        "User",
        "Date",
        "Direction",
        "From",
        "To",
        "Status",
        "Message",
        "Delivery Time",
        "Last Updated",
        "RingCentral ID"
    ]

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for column, header in enumerate(
        headers,
        start=1
    ):

        cell = sheet.cell(
            row=1,
            column=column,
            value=header
        )

        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # -----------------------------------------------------
    # DATA
    # -----------------------------------------------------

    for row_number, row in enumerate(
        rows,
        start=2
    ):

        values = [
            row["extension_number"],
            row["extension_name"],
            format_kyiv_time(row["creation_time"]),
            row["direction"],
            row["from_number"],
            row["to_number"],
            row["status"],
            row["message"],
            format_kyiv_time(row["delivery_time"]),
            format_kyiv_time(row["last_updated"]),
            row["ringcentral_id"]
        ]

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = sheet.cell(
                row=row_number,
                column=column,
                value=value
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(column == 8)
            )

    # -----------------------------------------------------
    # COLUMN WIDTHS
    # -----------------------------------------------------

    widths = {
        1: 14,
        2: 28,
        3: 22,
        4: 14,
        5: 20,
        6: 20,
        7: 20,
        8: 60,
        9: 22,
        10: 22,
        11: 25
    }

    for column, width in widths.items():

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = width

    # -----------------------------------------------------
    # FREEZE
    # -----------------------------------------------------

    sheet.freeze_panes = "A2"

    # -----------------------------------------------------
    # FILTER
    # -----------------------------------------------------

    sheet.auto_filter.ref = (
        f"A1:K{len(rows) + 1}"
    )

    # -----------------------------------------------------
    # SAVE TO MEMORY
    # -----------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# =========================================================
# OLD INTERACTIVE EXPORT
# =========================================================

def export_messages(extension):

    extension_id = extension["extension_id"]

    extension_number = (
        extension["extension_number"]
        or str(extension_id)
    )

    name = extension["name"] or "Unknown"

    print()
    print("=" * 60)
    print("EXPORT")
    print("=" * 60)

    print(
        f"Extension: {extension_number}"
    )

    print(
        f"User: {name}"
    )

    total = get_message_count(
        extension_id
    )

    print(
        f"Messages: {total}"
    )

    if total == 0:

        print()
        print("У этого extension нет сообщений.")

        return

    # -----------------------------------------------------
    # DATE FILTER
    # -----------------------------------------------------

    print()
    print("Фильтр по дате:")
    print("1. Все сообщения")
    print("2. Указать период")

    date_choice = input(
        "\nВыбор: "
    ).strip()

    date_from = None
    date_to = None

    if date_choice == "2":

        date_from = input(
            "Дата от (YYYY-MM-DD): "
        ).strip()

        date_to = input(
            "Дата до (YYYY-MM-DD): "
        ).strip()

        try:

            datetime.strptime(
                date_from,
                "%Y-%m-%d"
            )

            datetime.strptime(
                date_to,
                "%Y-%m-%d"
            )

        except ValueError:

            print(
                "Неверный формат даты."
            )

            return

    # -----------------------------------------------------
    # GET DATA
    # -----------------------------------------------------

    rows = get_messages_by_extensions(
        [extension_id],
        date_from=date_from,
        date_to=date_to
    )

    print()
    print(
        f"Подготовлено сообщений: {len(rows)}"
    )

    if not rows:

        print(
            "За выбранный период сообщений нет."
        )

        return

    # -----------------------------------------------------
    # EXCEL
    # -----------------------------------------------------

    workbook_file = generate_excel_from_messages(
        rows,
        title="SMS"
    )

    safe_name = "".join(
        c if c.isalnum() or c in (
            " ",
            "_",
            "-"
        )
        else "_"
        for c in name
    ).strip()

    if not safe_name:
        safe_name = "User"

    date_suffix = datetime.now().strftime(
        "%Y-%m-%d_%H-%M"
    )

    filename = (
        f"SMS_{extension_number}_"
        f"{safe_name}_"
        f"{date_suffix}.xlsx"
    )

    output_path = os.path.join(
        os.getcwd(),
        filename
    )

    with open(
        output_path,
        "wb"
    ) as file:

        file.write(
            workbook_file.getvalue()
        )

    print()
    print("=" * 60)
    print("ГОТОВО")
    print("=" * 60)

    print(
        f"Файл: {output_path}"
    )

    print(
        f"Сообщений: {len(rows)}"
    )

    print("=" * 60)


# =========================================================
# MAIN
# =========================================================

def main():

    print("=" * 60)
    print("SMS MONITOR — EXCEL EXPORT")
    print("=" * 60)

    extensions = get_extensions()

    if not extensions:

        print()
        print(
            "В БД нет активных User extensions."
        )

        return

    print()
    print("Доступные пользователи:")
    print()

    for index, extension in enumerate(
        extensions,
        start=1
    ):

        extension_number = (
            extension["extension_number"]
            or "-"
        )

        name = (
            extension["name"]
            or "Без имени"
        )

        count = get_message_count(
            extension["extension_id"]
        )

        print(
            f"{index:3}. "
            f"Ext. {extension_number:<10} "
            f"{name:<30} "
            f"{count} SMS"
        )

    print()

    choice = input(
        "Выберите extension: "
    ).strip()

    try:

        index = int(choice)

    except ValueError:

        print(
            "Введите номер из списка."
        )

        return

    if index < 1 or index > len(extensions):

        print(
            "Такого extension нет."
        )

        return

    extension = extensions[
        index - 1
    ]

    export_messages(
        extension
    )


if __name__ == "__main__":
    main()